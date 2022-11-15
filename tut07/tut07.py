
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill
from datetime import datetime
start_time = datetime.now()

#Help
def octant_analysis(mod=5000):
	pass

##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename. 

###Code
	#importing libraries
	import pandas as pd
	import glob
	import os


	try:

		path = "input" 
		filenames = glob.glob(path + "\*.xlsx")

		for file in filenames:

			print(str(file)[5:])

			wb = Workbook()
			ws = wb.active
			#ws.title = "output"

			d = pd.read_excel(file)


			d_1 = d["T"]
			d_2 = d["U"]
			d_3 = d["V"]
			d_4 = d["W"]
			d_5= {"+++": 1, "++-": -1, "-++": 2, "-+-": -2,
        		"--+": 3, "---": -3, "+-+": 4, "+--": -4}
			
			ini = round(d["U"].mean(),3)
			mi = round(d["V"].mean(),3)
			las = round(d["W"].mean(),3)
			o1 = [ini]
			o2 = [mi]
			o3 = [las]
			
			for i in range(len(d["T"])-1):
				o1.append(None)
				o2.append(None)
				o3.append(None)
			
			u1 = [round(i-ini,3) for i in d["U"]]
			v1 = [round(i-mi,3) for i in d["V"]]
			w1 = [round(i-las,3) for i in d["W"]]
			
			Octant = []
			for i in range(len(u1)):
				s = ""
				if(u1[i]<0):
					s += "-"
				else:
					s += "+"
				if(v1[i]<0):
					s += "-"
				else:
					s += "+"
				if(w1[i]<0):
					s += "-"
				else:
					s += "+"
				Octant.append(d_5[s])
				
			col = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
			text = ["T", "U", "V", "W", "U Avg", "V Avg", "W Avg", "U'=U-U Avg", "V'=V-V Avg", "W'=W-W Avg", "Octant"]
			ri = 2
			wh = [d_1, d_2, d_3, d_4, o1, o2, o3, u1, v1, w1, Octant]
			

			for i in range(len(col)):
				ws[f"{col[i]}{ri}"].value = text[i]
				for r in range(3,len(d_1)+3):
					ws[f"{col[i]}{r}"].value = wh[i][r-3]
			c= "M"
			ws[f"{c}{4}"].value = "Mod " + str(mod)

			c= "N"
			ws[f"{c}{1}"].value = "Overall Octant Count"

			thin = Side(border_style="thin", color="000000")

			co = ["N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF"]
			tex = ["Octant ID", "+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4", "Rank Octant 1", "Rank Octant -1", "Rank Octant 2", "Rank Octant -2", "Rank Octant 3", "Rank Octant -3", "Rank Octant 4", "Rank Octant -4", "Rank1 Octant ID", "Rank1 Octant Name"]

			for i in range(len(co)):
				ws[f"{co[i]}{3}"].value = tex[i]
				ws[f"{co[i]}{3}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

			for p in range(len(co)):

				ws[f"{co[p]}{4}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
				OctantName = ["Internal outward interaction", "External outward interaction", "External Ejection", "Internal Ejection", "External inward interaction", "Internal inward interaction", "Internal sweep", "External sweep"]
				rr = [Octant.count(1), Octant.count(-1), Octant.count(2), Octant.count(-2), Octant.count(3), Octant.count(-3), Octant.count(4), Octant.count(-4)]
				rrr = [Octant.count(1), Octant.count(-1), Octant.count(2), Octant.count(-2), Octant.count(3), Octant.count(-3), Octant.count(4), Octant.count(-4)]
				rrr.sort(reverse=True)
				r1=r2=r3=r4=r5=r6=r7=r8=c=0
				for i in range(8):
					c=0
					for j in range(8):
						if(rr[i]==rrr[j]):
							if(j==0):
								if(i==0):
									ro=1
								elif(i==1):
									ro=-1
								elif(i==2):
									ro=2
								elif(i==3):
									ro=-2
								elif(i==4):
									ro=3
								elif(i==5):
									ro=-3
								elif(i==6):
									ro=4
								elif(i==7):
									ro=-4
								rn=OctantName[i]
							if(i==0):
								r1=(j+1)
							elif(i==1):
								r2=(j+1)
							elif(i==2):
								r3=(j+1)
							elif(i==3):
								r4=(j+1)
							elif(i==4):
								r5=(j+1)
							elif(i==5):
								r6=(j+1)
							elif(i==6):
								r7=(j+1)
							elif(i==7):
								r8=(j+1)
							c=1
						if(c==1):
							break
				if(co[p]=="N"):
					ws[f"{co[p]}{4}"].value = "Overall Count"
				elif(co[p]=="O"):
					ws[f"{co[p]}{4}"].value = Octant.count(1)
				elif(co[p]=="P"):
					ws[f"{co[p]}{4}"].value = Octant.count(-1)
				elif(co[p]=="Q"):
					ws[f"{co[p]}{4}"].value = Octant.count(2)
				elif(co[p]=="R"):
					ws[f"{co[p]}{4}"].value = Octant.count(-2)
				elif(co[p]=="S"):
					ws[f"{co[p]}{4}"].value = Octant.count(3)
				elif(co[p]=="T"):
					ws[f"{co[p]}{4}"].value = Octant.count(-3)
				elif(co[p]=="U"):
					ws[f"{co[p]}{4}"].value = Octant.count(4)
				elif(co[p]=="V"):
					ws[f"{co[p]}{4}"].value = Octant.count(-4)
				elif(co[p]=="W"):
					ws[f"{co[p]}{4}"].value = r1
					if(r1==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="X"):
					ws[f"{co[p]}{4}"].value = r2
					if(r2==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="Y"):
					ws[f"{co[p]}{4}"].value = r3
					if(r3==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="Z"):
					ws[f"{co[p]}{4}"].value = r4
					if(r4==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="AA"):
					ws[f"{co[p]}{4}"].value = r5
					if(r5==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="AB"):
					ws[f"{co[p]}{4}"].value = r6
					if(r6==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="AC"):
					ws[f"{co[p]}{4}"].value = r7
					if(r7==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="AD"):
					ws[f"{co[p]}{4}"].value = r8
					if(r8==1):
						ws[f"{co[p]}{4}"].fill = PatternFill("solid", fgColor="00FFFF00")
				elif(co[p]=="AE"):
					ws[f"{co[p]}{4}"].value = ro
				elif(co[p]=="AF"):
					ws[f"{co[p]}{4}"].value = rn

			b=4
			ofc = []
			for k in range(0, len(d["T"]), mod):
				b+=1
				for p in range(len(co)):

					ws[f"{co[p]}{b}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
					OctantName = ["Internal outward interaction", "External outward interaction", "External Ejection", "Internal Ejection", "External inward interaction", "Internal inward interaction", "Internal sweep", "External sweep"]
					rr = [Octant[k:k+mod].count(1), Octant[k:k+mod].count(-1), Octant[k:k+mod].count(2), Octant[k:k+mod].count(-2), Octant[k:k+mod].count(3), Octant[k:k+mod].count(-3), Octant[k:k+mod].count(4), Octant[k:k+mod].count(-4)]
					rrr = [Octant[k:k+mod].count(1), Octant[k:k+mod].count(-1), Octant[k:k+mod].count(2), Octant[k:k+mod].count(-2), Octant[k:k+mod].count(3), Octant[k:k+mod].count(-3), Octant[k:k+mod].count(4), Octant[k:k+mod].count(-4)]
					rrr.sort(reverse=True)
					r1=r2=r3=r4=r5=r6=r7=r8=c=0
					for i in range(8):
						c=0
						for j in range(8):
							if(rr[i]==rrr[j]):
								if(j==0):
									if(i==0):
										ro=1
									elif(i==1):
										ro=-1
									elif(i==2):
										ro=2
									elif(i==3):
										ro=-2
									elif(i==4):
										ro=3
									elif(i==5):
										ro=-3
									elif(i==6):
										ro=4
									elif(i==7):
										ro=-4
									rn=OctantName[i]
									
								if(i==0):
									r1=(j+1)
								elif(i==1):
									r2=(j+1)
								elif(i==2):
									r3=(j+1)
								elif(i==3):
									r4=(j+1)
								elif(i==4):
									r5=(j+1)
								elif(i==5):
									r6=(j+1)
								elif(i==6):
									r7=(j+1)
								elif(i==7):
									r8=(j+1)
								c=1
							if(c==1):
								break
					if(co[p]=="N"):
						ws[f"{co[p]}{b}"].value = str(k) + "-" + str(k+mod-1)
					elif(co[p]=="O"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(1)
					elif(co[p]=="P"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(-1)
					elif(co[p]=="Q"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(2)
					elif(co[p]=="R"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(-2)
					elif(co[p]=="S"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(3)
					elif(co[p]=="T"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(-3)
					elif(co[p]=="U"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(4)
					elif(co[p]=="V"):
						ws[f"{co[p]}{b}"].value = Octant[k:k+mod].count(-4)
					elif(co[p]=="W"):
						ws[f"{co[p]}{b}"].value = r1
						if(r1==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="X"):
						ws[f"{co[p]}{b}"].value = r2
						if(r2==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="Y"):
						ws[f"{co[p]}{b}"].value = r3
						if(r3==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="Z"):
						ws[f"{co[p]}{b}"].value = r4
						if(r4==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="AA"):
						ws[f"{co[p]}{b}"].value = r5
						if(r5==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="AB"):
						ws[f"{co[p]}{b}"].value = r6
						if(r6==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="AC"):
						ws[f"{co[p]}{b}"].value = r7
						if(r7==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="AD"):
						ws[f"{co[p]}{b}"].value = r8
						if(r8==1):
							ws[f"{co[p]}{b}"].fill = PatternFill("solid", fgColor="00FFFF00")
					elif(co[p]=="AE"):
						ws[f"{co[p]}{b}"].value = ro
					elif(co[p]=="AF"):
						ws[f"{co[p]}{b}"].value = rn
						ofc.append(rn)

			b+=2
			g = ["AC", "AD", "AE"]
			ws[f"{g[0]}{b}"].value = "Octant ID"
			ws[f"{g[1]}{b}"].value = "Octant Name"
			ws[f"{g[2]}{b}"].value = "Count of Rank 1 Mod Values"
			ws[f"{g[0]}{b}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
			ws[f"{g[1]}{b}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
			ws[f"{g[2]}{b}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
			b+=1

			oo = [1, -1, 2, -2, 3, -3, 4, -4]
			for i in range(8):
				ws[f"{g[0]}{b+i}"].value = oo[i]
				ws[f"{g[1]}{b+i}"].value = OctantName[i]
				ws[f"{g[2]}{b+i}"].value = ofc.count(OctantName[i])
				ws[f"{g[0]}{b+i}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
				ws[f"{g[1]}{b+i}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
				ws[f"{g[2]}{b+i}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)


			cc = "AH"
			ws[f"{cc}{4}"].value = "From"
			b = 18
			for i in range(0, len(d["T"]), mod):
				ws[f"{cc}{b}"].value = "From"
				b+=13
			
			transition_1 = []
			transition_11 = []
			transition_2 = []
			transition_22 = []
			transition_3 = []
			transition_33 = []
			transition_4 = []
			transition_44 = []
			
			for i in range(0, (len(d["T"]))-1, 1):
				if(Octant[i+1]==+1):
					transition_1.append(Octant[i])
				elif(Octant[i+1]==-1):
					transition_11.append(Octant[i])
				elif(Octant[i+1]==+2):
					transition_2.append(Octant[i])
				elif(Octant[i+1]==-2):
					transition_22.append(Octant[i])
				elif(Octant[i+1]==+3):
					transition_3.append(Octant[i])
				elif(Octant[i+1]==-3):
					transition_33.append(Octant[i])
				elif(Octant[i+1]==+4):
					transition_4.append(Octant[i])
				elif(Octant[i+1]==-4):
					transition_44.append(Octant[i])

			cc = "AI"
			ws[f"{cc}{1}"].value = "Overall Transition Count"
			cc = "AJ"
			ws[f"{cc}{2}"].value = "To"

			gg = ["AI", "AJ", "AK", "AL", "AM", "AN", "AO", "AP", "AQ"]

			for i in range(len(gg)):
				if(i==0):
					ws[f"{gg[i]}{3}"].value = "Octant #"
				else:
					ws[f"{gg[i]}{3}"].value = oo[i-1]
				ws[f"{gg[i]}{3}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)

			b=4
			for i in range(8):
				ma = -1
				rc = 0
				cr =0
				for j in range(len(gg)):
					if(j==0):
						ws[f"{gg[j]}{b+i}"].value = oo[i]
					elif(j==1):
						ws[f"{gg[j]}{b+i}"].value = transition_1.count(oo[i])
						if(ma<transition_1.count(oo[i])):
							ma = transition_1.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==2):
						ws[f"{gg[j]}{b+i}"].value = transition_11.count(oo[i])
						if(ma<transition_11.count(oo[i])):
							ma = transition_11.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==3):
						ws[f"{gg[j]}{b+i}"].value = transition_2.count(oo[i])
						if(ma<transition_2.count(oo[i])):
							ma = transition_2.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==4):
						ws[f"{gg[j]}{b+i}"].value = transition_22.count(oo[i])
						if(ma<transition_22.count(oo[i])):
							ma = transition_22.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==5):
						ws[f"{gg[j]}{b+i}"].value = transition_3.count(oo[i])
						if(ma<transition_3.count(oo[i])):
							ma = transition_3.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==6):
						ws[f"{gg[j]}{b+i}"].value = transition_33.count(oo[i])
						if(ma<transition_33.count(oo[i])):
							ma = transition_33.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==7):
						ws[f"{gg[j]}{b+i}"].value = transition_4.count(oo[i])
						if(ma<transition_4.count(oo[i])):
							ma = transition_4.count(oo[i])
							rc = b+i
							cr = gg[j]
					elif(j==8):
						ws[f"{gg[j]}{b+i}"].value = transition_44.count(oo[i])
						if(ma<transition_44.count(oo[i])):
							ma = transition_44.count(oo[i])
							rc = b+i
							cr = gg[j]
					ws[f"{gg[j]}{b+i}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
				ws[f"{cr}{rc}"].fill = PatternFill("solid", fgColor="00FFFF00")
			b=15
			for ti in range(0, len(d["T"]), mod):
				ws[f"{gg[0]}{b}"].value = "Mod Transition Count"
				b+=1
				ws[f"{gg[0]}{b}"].value = str(ti)+"-"+str(ti+mod-1)
				ws[f"{gg[1]}{b}"].value = "To"
				b+=1
				for j in range(len(gg)):
					if(j==0):
						ws[f"{gg[j]}{b}"].value = "Octant #"
					else:
						ws[f"{gg[j]}{b}"].value = oo[j-1]
					ws[f"{gg[j]}{b}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
				transition_01 = []
				transition_011 = []
				transition_02 = []
				transition_022 = []
				transition_03 = []
				transition_033 = []
				transition_04 = []
				transition_044 = []
			
				for i in range(ti, ti+mod, 1):
					if(i<len(d["T"])-1):
						if(Octant[i+1]==+1):
							transition_1.append(Octant[i])
						elif(Octant[i+1]==-1):
							transition_11.append(Octant[i])
						elif(Octant[i+1]==+2):
							transition_2.append(Octant[i])
						elif(Octant[i+1]==-2):
							transition_22.append(Octant[i])
						elif(Octant[i+1]==+3):
							transition_3.append(Octant[i])
						elif(Octant[i+1]==-3):
							transition_33.append(Octant[i])
						elif(Octant[i+1]==+4):
							transition_4.append(Octant[i])
						elif(Octant[i+1]==-4):
							transition_44.append(Octant[i])

				b+=1
				for i in range(8):
					ma = -1
					rc = 0
					cr =0
					for j in range(len(gg)):
						if(j==0):
							ws[f"{gg[j]}{b+i}"].value = oo[i]
						elif(j==1):
							ws[f"{gg[j]}{b+i}"].value = transition_1.count(oo[i])
							if(ma<transition_1.count(oo[i])):
								ma = transition_1.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==2):
							ws[f"{gg[j]}{b+i}"].value = transition_11.count(oo[i])
							if(ma<transition_11.count(oo[i])):
								ma = transition_11.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==3):
							ws[f"{gg[j]}{b+i}"].value = transition_2.count(oo[i])
							if(ma<transition_2.count(oo[i])):
								ma = transition_2.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==4):
							ws[f"{gg[j]}{b+i}"].value = transition_22.count(oo[i])
							if(ma<transition_22.count(oo[i])):
								ma = transition_22.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==5):
							ws[f"{gg[j]}{b+i}"].value = transition_3.count(oo[i])
							if(ma<transition_3.count(oo[i])):
								ma = transition_3.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==6):
							ws[f"{gg[j]}{b+i}"].value = transition_33.count(oo[i])
							if(ma<transition_33.count(oo[i])):
								ma = transition_33.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==7):
							ws[f"{gg[j]}{b+i}"].value = transition_4.count(oo[i])
							if(ma<transition_4.count(oo[i])):
								ma = transition_4.count(oo[i])
								rc = b+i
								cr = gg[j]
						elif(j==8):
							ws[f"{gg[j]}{b+i}"].value = transition_44.count(oo[i])
							if(ma<transition_44.count(oo[i])):
								ma = transition_44.count(oo[i])
								rc = b+i
								cr = gg[j]
						ws[f"{gg[j]}{b+i}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
					ws[f"{cr}{rc}"].fill = PatternFill("solid", fgColor="00FFFF00")
				b+=10



			
	except:
	    print("Something went wrong while opening the file or file is not found.")
        


from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_analysis(mod)






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
